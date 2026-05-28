"""
export_utils.py — Excel and PDF export.
Returns raw bytes for Flask send_file().
"""
import io
from datetime import datetime


def generate_excel(result: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Poultry Data"

    GD, GM, GL = "1B5E20", "2A6B2A", "EBF4EB"
    AM, BL, WH = "FFF8E1", "E3F2FD", "FFFFFF"
    BC = "BDBDBD"

    def hf(sz=10, b=True, c=WH):
        return Font(name="Calibri", size=sz, bold=b, color=c)
    def bf(sz=10, b=False, c="212121"):
        return Font(name="Calibri", size=sz, bold=b, color=c)
    def fl(c): return PatternFill("solid", fgColor=c)
    def bd():
        s = Side(style="thin", color=BC)
        return Border(left=s, right=s, top=s, bottom=s)
    def ct(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def lt(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def wc(r, c, v, font=None, fill=None, align=None, bdr=None):
        cell = ws.cell(row=r, column=c, value=v)
        if font:  cell.font      = font
        if fill:  cell.fill      = fill
        if align: cell.alignment = align
        if bdr:   cell.border    = bdr
        return cell

    def mr(r, c1, c2, v, font=None, fill=None, align=None, h=None):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws.cell(row=r, column=c1, value=v)
        if font:  cell.font      = font
        if fill:  cell.fill      = fill
        if align: cell.alignment = align
        if h:     ws.row_dimensions[r].height = h

    # Title
    mr(1, 1, 6, "Poultry Record Digitizer — Extracted Data",
       font=hf(14), fill=fl(GD), align=ct(), h=36)
    mr(2, 1, 6,
       f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}  |  "
       f"OCR: {result.get('ocr_engine', 'N/A')}",
       font=bf(9, c="555555"), fill=fl(GL), align=ct(), h=20)
    ws.append([])

    # Key metrics
    mr(4, 1, 6, "KEY METRICS",
       font=hf(10, c=GD), fill=fl(GL), align=lt(), h=22)
    for col, h in enumerate(["Metric", "Value"], 1):
        wc(5, col, h, font=hf(), fill=fl(GM), align=ct(), bdr=bd())
    ws.row_dimensions[5].height = 22

    metrics = [
        ("Feed (kg/gms)",        result.get("feed")),
        ("Eggs",                 result.get("eggs")),
        ("Mortality",            result.get("mortality")),
        ("Birds",                result.get("birds")),
        ("Date",                 result.get("date")),
        ("Batch",                result.get("batch")),
        ("Latest ABW (gms)",     result.get("latest_abw")),
        ("Latest FCR",           result.get("latest_fcr")),
        ("Total Mortality",      result.get("total_mortality")),
        ("ABW Readings Found",   len(result.get("abw_values", []))),
        ("FCR Readings Found",   len(result.get("fcr_values", []))),
    ]
    for i, (lbl, val) in enumerate(metrics):
        r  = 6 + i
        bg = AM if i % 2 == 0 else WH
        wc(r, 1, lbl,
           font=bf(b=True), fill=fl(bg), align=lt(), bdr=bd())
        wc(r, 2, val if val is not None else "—",
           font=bf(), fill=fl(bg), align=ct(), bdr=bd())
        ws.row_dimensions[r].height = 20

    # ABW table
    abw = result.get("abw_values", [])
    r   = 6 + len(metrics) + 2
    if abw:
        mr(r, 1, 6, "ABW READINGS (Actual Body Weight — grams)",
           font=hf(10, c=GD), fill=fl(GL), align=lt(), h=22)
        r += 1
        for col, h in enumerate(["#","ABW (gms)","Est. Week","Change","% Change",""], 1):
            wc(r, col, h, font=hf(), fill=fl(GM), align=ct(), bdr=bd())
        ws.row_dimensions[r].height = 22
        r += 1
        for i, val in enumerate(abw):
            prev = abw[i-1] if i > 0 else None
            chg  = (val - prev) if prev is not None else None
            pct  = round((chg / prev) * 100, 1) if chg is not None and prev else None
            bg   = BL if i % 2 == 0 else WH
            for col, v in enumerate(
                [i+1, val, f"~Week {i+1}",
                 chg if chg is not None else "—",
                 f"{pct}%" if pct is not None else "—", ""], 1
            ):
                wc(r, col, v, font=bf(), fill=fl(bg), align=ct(), bdr=bd())
            ws.row_dimensions[r].height = 20
            r += 1
        r += 1

    # FCR table
    fcr = result.get("fcr_values", [])
    if fcr:
        mr(r, 1, 6, "FCR READINGS (Feed Conversion Ratio)",
           font=hf(10, c=GD), fill=fl(GL), align=lt(), h=22)
        r += 1
        for col, h in enumerate(["#","FCR","Rating","Interpretation","",""], 1):
            wc(r, col, h, font=hf(), fill=fl(GM), align=ct(), bdr=bd())
        ws.row_dimensions[r].height = 22
        r += 1
        for i, val in enumerate(fcr):
            if val < 0.8:   rt, rc, interp = "Excellent", "00695C", "Very efficient"
            elif val < 1.0: rt, rc, interp = "Good",      "2A6B2A", "Normal pattern"
            elif val < 1.2: rt, rc, interp = "Average",   "D97706", "Efficiency declining"
            else:           rt, rc, interp = "Poor",      "DC2626", "High feed vs gain"
            bg = AM if i % 2 == 0 else WH
            for col, v in enumerate([i+1, val, rt, interp, "", ""], 1):
                fc = rc if col == 3 else "212121"
                wc(r, col, v,
                   font=bf(b=(col==3), c=fc), fill=fl(bg),
                   align=ct() if col != 4 else lt(), bdr=bd())
            ws.row_dimensions[r].height = 20
            r += 1
        r += 1

    # Medicine notes
    notes = result.get("medicine_notes", [])
    if notes:
        mr(r, 1, 6, "MEDICINE / VACCINE NOTES",
           font=hf(10, c=GD), fill=fl(GL), align=lt(), h=22)
        r += 1
        for col, h in enumerate(["#", "Note"], 1):
            wc(r, col, h, font=hf(), fill=fl(GM), align=ct(), bdr=bd())
        ws.row_dimensions[r].height = 22
        r += 1
        for i, note in enumerate(notes):
            bg = AM if i % 2 == 0 else WH
            wc(r, 1, i+1, font=bf(), fill=fl(bg), align=ct(), bdr=bd())
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            wc(r, 2, note, font=bf(), fill=fl(bg), align=lt(), bdr=bd())
            ws.row_dimensions[r].height = 20
            r += 1

    # Column widths
    for i, w in enumerate([6, 18, 18, 36, 14, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    # Sheet 2: Raw OCR
    ws2 = wb.create_sheet("Raw OCR")
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 90
    for col, h in enumerate(["Field", "Content"], 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = hf(); c.fill = fl(GM)
    for rn, (lbl, key) in enumerate(
        [("Cleaned", "cleaned_text"), ("Raw", "raw_text")], 2
    ):
        ws2.cell(row=rn, column=1, value=lbl).font = bf(b=True)
        c = ws2.cell(row=rn, column=2, value=result.get(key, ""))
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[rn].height = 90

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_pdf(result: dict) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer,
        Table, TableStyle, HRFlowable
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm,  bottomMargin=2*cm,
        title="Poultry Record"
    )

    CG  = colors.HexColor("#2A6B2A")
    CGD = colors.HexColor("#1B5E20")
    CGL = colors.HexColor("#EBF4EB")
    CA  = colors.HexColor("#FFF8E1")
    CBL = colors.HexColor("#E3F2FD")
    CGR = colors.HexColor("#757575")
    CBD = colors.HexColor("#BDBDBD")
    CW  = colors.white
    W   = A4[0] - 4*cm

    ts  = ParagraphStyle("t",  fontName="Helvetica-Bold", fontSize=15,
                          textColor=CW, alignment=1)
    h2  = ParagraphStyle("h2", fontName="Helvetica-Bold", fontSize=11,
                          textColor=CGD, spaceBefore=14, spaceAfter=6)
    sm  = ParagraphStyle("sm", fontName="Helvetica", fontSize=8,
                          textColor=CGR, leading=12)

    def tstyle(extra=None):
        base = [
            ("BACKGROUND",    (0,0), (-1,0),  CG),
            ("TEXTCOLOR",     (0,0), (-1,0),  CW),
            ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,-1), 9),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [CW, CA]),
            ("ALIGN",         (0,0), (-1,-1), "CENTER"),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING",    (0,0), (-1,-1), 6),
            ("BOTTOMPADDING", (0,0), (-1,-1), 6),
            ("GRID",          (0,0), (-1,-1), 0.4, CBD),
        ]
        if extra: base.extend(extra)
        return TableStyle(base)

    story = []

    # Title banner
    banner = Table(
        [[Paragraph("Poultry Record Digitizer — Data Report", ts)]],
        colWidths=[W]
    )
    banner.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), CGD),
        ("TOPPADDING",    (0,0), (-1,-1), 12),
        ("BOTTOMPADDING", (0,0), (-1,-1), 12),
    ]))
    story += [banner, Spacer(1, 6)]
    story.append(Paragraph(
        f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')} | "
        f"OCR: {result.get('ocr_engine', 'N/A')}", sm
    ))
    story += [Spacer(1,10), HRFlowable(width=W, thickness=0.5, color=CG),
              Spacer(1,10)]

    def v(k):
        val = result.get(k)
        return str(val) if val is not None else "—"

    # Key metrics
    story.append(Paragraph("Key Metrics", h2))
    mt = Table([
        ["Metric",              "Value"],
        ["Feed (kg/gms)",       v("feed")],
        ["Eggs",                v("eggs")],
        ["Mortality",           v("mortality")],
        ["Birds",               v("birds")],
        ["Date",                v("date")],
        ["Batch",               v("batch")],
        ["Latest ABW (gms)",    v("latest_abw")],
        ["Latest FCR",          v("latest_fcr")],
        ["ABW Readings Found",  str(len(result.get("abw_values", [])))],
        ["FCR Readings Found",  str(len(result.get("fcr_values", [])))],
    ], colWidths=[W*0.55, W*0.45])
    mt.setStyle(tstyle([
        ("FONTNAME", (0,1),(0,-1), "Helvetica-Bold"),
        ("ALIGN",    (1,0),(1,-1), "CENTER"),
    ]))
    story += [mt, Spacer(1, 14)]

    # ABW table
    abw = result.get("abw_values", [])
    if abw:
        story.append(Paragraph("ABW Readings (Actual Body Weight — grams)", h2))
        rows = [["#", "ABW (gms)", "Est. Week", "Change"]]
        for i, val in enumerate(abw):
            prev = abw[i-1] if i > 0 else None
            chg  = (f"+{val-prev} gms" if prev is not None and val >= prev
                    else (f"{val-prev} gms" if prev is not None else "—"))
            rows.append([str(i+1), str(val), f"~Week {i+1}", chg])
        at = Table(rows, colWidths=[W*.08, W*.22, W*.20, W*.50])
        at.setStyle(tstyle([("ROWBACKGROUNDS",(0,1),(-1,-1),[CW,CBL])]))
        story += [at, Spacer(1, 14)]

    # FCR table
    fcr = result.get("fcr_values", [])
    if fcr:
        story.append(Paragraph("FCR Readings (Feed Conversion Ratio)", h2))
        rows = [["#", "FCR", "Rating", "Interpretation"]]
        for i, val in enumerate(fcr):
            if val < 0.8:   rt, interp = "Excellent", "Very efficient growth"
            elif val < 1.0: rt, interp = "Good",      "Normal growth pattern"
            elif val < 1.2: rt, interp = "Average",   "Feed efficiency declining"
            else:           rt, interp = "Poor",      "High feed vs weight gain"
            rows.append([str(i+1), str(val), rt, interp])
        ft = Table(rows, colWidths=[W*.08, W*.15, W*.18, W*.59])
        ft.setStyle(tstyle([("ALIGN",(3,1),(3,-1),"LEFT")]))
        story += [ft, Spacer(1, 14)]

    # Medicine notes
    notes = result.get("medicine_notes", [])
    if notes:
        story.append(Paragraph("Medicine / Vaccine Notes", h2))
        rows = [["#", "Note"]] + [[str(i+1), n] for i, n in enumerate(notes)]
        nt = Table(rows, colWidths=[W*.08, W*.92])
        nt.setStyle(tstyle([("ALIGN",(1,1),(1,-1),"LEFT")]))
        story += [nt, Spacer(1, 14)]

    story += [
        Spacer(1, 20),
        HRFlowable(width=W, thickness=0.5, color=CGR),
        Spacer(1, 6),
        Paragraph("Generated by Poultry Record Digitizer — College Project", sm),
    ]

    doc.build(story)
    buf.seek(0)
    return buf.read()